import os
import re
import openpyxl
import pandas as pd


def merge_cells_to_values(file_path):
    """
    处理合并单元格，将合并的单元格值填充到所有合并的单元格内。
    """
    wb = openpyxl.load_workbook(file_path)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for merged_range in ws.merged_cells.ranges:
            min_row, min_col, max_row, max_col = merged_range.bounds
            cell_value = ws.cell(row=min_row, column=min_col).value
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col, value=cell_value)

    wb.save(file_path)


def flatten_headers(df):
    """
    清理并扁平化DataFrame的多层表头，生成一个连续、无冗余的表头。
    """
    def clean_header(header):
        # 提取多层表头中的有效信息，忽略未命名的列
        levels = [str(i).strip() for i in header if
                  i and isinstance(i, str) and 'Unnamed' not in str(i) and 'nan' not in str(i)]
        return " - ".join(levels) if levels else None

    # 生成清洗后的列名，若为空则给定默认名称
    cleaned_columns = [clean_header(col) or f"未命名列_{i}" for i, col in enumerate(df.columns.values)]
    # 确保列名唯一性，移除重复的无意义列
    unique_columns = []
    for col in cleaned_columns:
        if col.startswith("未命名列") and col in unique_columns:
            continue
        unique_columns.append(col)
    df.columns = unique_columns
    return df


def detect_header_rows(sheet_data, max_header_rows=5):
    """
    检测表头行位置，根据文本单元格占比确定表头行。最大检测行数限制为max_header_rows，以避免误判。
    """
    header_rows = []
    for i, row in sheet_data.iterrows():
        text_count = sum(isinstance(cell, str) for cell in row)
        non_empty_count = sum(pd.notna(cell) for cell in row)

        if non_empty_count > 0 and text_count / non_empty_count > 0.5:
            header_rows.append(i)
        elif header_rows:  # 如果已找到表头行，遇到不符合条件的行则停止
            break

        if len(header_rows) >= max_header_rows:
            break

    return header_rows if header_rows and max(header_rows) < len(sheet_data) else [0,1,2]


def flatten_and_combine_headers(df):
    """
    合并多层表头行，按列拼接字符串生成一个完整的标题行。
    """
    headers = []
    for i, col in enumerate(df.columns.values):
        combined_header = " - ".join([str(part).strip() for part in col if part and 'Unnamed' not in str(part)])
        headers.append(combined_header if combined_header else f"未命名列_{i}")
    df.columns = headers
    return df


def split_sheet_if_parallel(sheet_data):
    """
    检测表格中的并行结构，并将其拆分为左右两部分（即若表格包含左右结构）。
    """
    for col in sheet_data.columns:
        text_ratio = sheet_data[col].apply(lambda x: isinstance(x, str)).mean()
        if text_ratio >= 0.6:
            left_part = sheet_data.iloc[:, :sheet_data.columns.get_loc(col)]
            right_part = sheet_data.iloc[:, sheet_data.columns.get_loc(col):]
            return left_part, right_part
    return sheet_data, None


def process_table(table_data, sheet_name, key_value_list):
    """
    处理单个表格数据，提取并存储键值对信息。
    """
    chinese_char_pattern = re.compile(r'[\u4e00-\u9fa5]')

    for _, row in table_data.iterrows():
        row = row.dropna()
        row_label = None  # 初始化行标签

        for col_name, value in row.items():
            if row_label is None and isinstance(value, str):
                row_label = value.strip()
            elif isinstance(value, (int, float)) and pd.notna(value) and row_label:
                if chinese_char_pattern.search(col_name):
                    key = f"{row_label.strip()} {col_name.strip()}".replace("\n", "") \
                        .replace("nan", "").replace(" ", "").replace("未命名列", "").replace("VALID#", "") \
                        .replace("AD_NAME#", "").replace("AD_BJ#", "").replace("-", "")
                    key_value_list.append(f"{key}：{value}")


def parse_excel_to_key_value_list(file_path):
    """
    解析Excel文件中的表格，并提取键值对信息。
    """
    file_extension = os.path.splitext(file_path)[1].lower()
    # 根据扩展名自动选择引擎
    if file_extension == '.xlsx':
        engine = 'openpyxl'  # 使用 openpyxl 引擎读取 .xlsx 文件
    elif file_extension == '.xls':
        engine = 'xlrd'  # 使用 xlrd 引擎读取 .xls 文件
    try:
        sheets = pd.read_excel(file_path, engine=engine, sheet_name=None)
    except Exception as e:
        raise Exception(f"读取 Excel 文件时发生错误: {e}")
    key_value_list = []
    for sheet_name, sheet_data in sheets.items():
        if sheet_name in ["封面", "目录"]:
            continue
        sheet_data.dropna(how='all', axis=0, inplace=True)
        sheet_data.dropna(how='all', axis=1, inplace=True)

        # 检测表头行并进行处理
        header_rows = detect_header_rows(sheet_data)

        if header_rows:
            if max(header_rows) < len(sheet_data):
                sheet_data.columns = pd.MultiIndex.from_arrays(sheet_data.iloc[header_rows].values)
                sheet_data = sheet_data.drop(header_rows).reset_index(drop=True)
                sheet_data = flatten_and_combine_headers(sheet_data)  # 合并标题行

            process_table(sheet_data, sheet_name, key_value_list)

    return key_value_list


def unfreeze_panes(file_path, output_path=None):
    """
    取消 Excel 文件的冻结窗格。
    """
    wb = openpyxl.load_workbook(file_path)

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        if ws.freeze_panes:
            ws.freeze_panes = None
    if output_path:
        wb.save(output_path)
    else:
        wb.save(file_path)


def process_tableFiles_in_folder(folder_path):
    from parser.folder_read import add_tags_to_sentences
    key_value_pairs = []
    file_list = [os.path.join(folder_path, file) for file in os.listdir(folder_path)
                 if os.path.isfile(os.path.join(folder_path, file)) and file.lower().endswith(('.xls', '.xlsx'))]
    for file_path in file_list:
        try:
            key_value_pair = parse_excel_to_key_value_list(file_path)
            if "VALID#" in "".join(key_value_pair):
                unfreeze_panes(file_path)
                key_value_pair = parse_excel_to_key_value_list(file_path)
            key_value_pairs = key_value_pairs + key_value_pair
        except Exception as e:
            print(f"处理文件 {file_path} 时发生错误: {e}")
            continue
    return add_tags_to_sentences(list(set(key_value_pairs)))

